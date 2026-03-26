import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import json
from datetime import datetime

from app.ml.predict import predict_emissions_with_uncertainty  # FIXED: Changed import
from app.models.process_data import ProcessData

class FileProcessor:
    """Process uploaded Excel/CSV files into process data batches"""
    
    # Column mappings for common file formats
    COLUMN_MAPPINGS = {
        # Excel template columns → database fields
        "batch_id": ["Batch ID", "BatchID", "Batch", "ID"],
        "raw_material_type": ["Raw Material Type", "Material Type", "Raw Material", "RM Type"],
        "material_type": ["Material Type", "Product Type", "Metal Type", "Material"],
        "energy_source_type": ["Energy Source", "Energy Type", "Fuel Type", "Source Type"],
        "raw_material_quantity": ["Raw Material Qty", "RM Quantity", "Virgin Material", "Raw Qty"],
        "recycled_material_quantity": ["Recycled Material", "Recycled Qty", "Scrap Input", "Recycled"],
        "energy_consumption_kwh": ["Energy Consumption", "Energy kWh", "Power Use", "Energy"],
        "water_consumption_liters": ["Water Consumption", "Water Use", "Water L", "Water"],
        "production_volume": ["Production Volume", "Output", "Product Qty", "Volume"],
        "ore_grade": ["Ore Grade", "Grade", "Concentration", "Quality"],
        "waste_slag_quantity": ["Waste Slag", "Slag Qty", "Waste", "Residue"],
        "scrap_content_percentage": ["Scrap Content", "Scrap %", "Recycled %", "Scrap"],
        "recycling_rate_percentage": ["Recycling Rate", "Recycle %", "Recovery Rate", "Recycling"]
    }
    
    @classmethod
    def detect_file_format(cls, file_path: str) -> str:
        """Detect file format and structure"""
        try:
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, nrows=5)
            elif file_path.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file_path, nrows=5)
            else:
                return "unknown"
            
            # Check for known column patterns
            columns = [str(col).lower() for col in df.columns]
            
            # Check for EcoMetal template
            template_columns = ['batch id', 'raw material type', 'energy consumption']
            if any(template in ' '.join(columns) for template in template_columns):
                return "ecometal_template"
            
            # Check for generic LCA format
            lca_keywords = ['co2', 'emission', 'energy', 'material', 'water']
            if any(keyword in ' '.join(columns) for keyword in lca_keywords):
                return "generic_lca"
            
            return "unknown"
            
        except Exception as e:
            print(f"Format detection error: {str(e)}")
            return "unknown"
    
    @classmethod
    def map_columns(cls, df_columns: List[str]) -> Dict[str, str]:
        """Map uploaded file columns to database fields"""
        mapping = {}
        for db_field, possible_names in cls.COLUMN_MAPPINGS.items():
            for uploaded_col in df_columns:
                if isinstance(uploaded_col, str):
                    # Check for exact match or partial match
                    for possible_name in possible_names:
                        if possible_name.lower() in uploaded_col.lower() or uploaded_col.lower() in possible_name.lower():
                            mapping[uploaded_col] = db_field
                            break
        
        return mapping
    
    @classmethod
    def validate_batch_data(cls, data: Dict) -> List[str]:
        """Validate batch data and return error messages"""
        errors = []
        
        # Required fields
        required_fields = ['batch_id', 'raw_material_quantity', 'energy_consumption_kwh', 'production_volume']
        for field in required_fields:
            if field not in data or data[field] is None:
                errors.append(f"Missing required field: {field}")
        
        # Numeric validation
        numeric_fields = [
            'raw_material_quantity', 'recycled_material_quantity',
            'energy_consumption_kwh', 'water_consumption_liters',
            'production_volume', 'ore_grade', 'waste_slag_quantity',
            'scrap_content_percentage', 'recycling_rate_percentage'
        ]
        
        for field in numeric_fields:
            if field in data and data[field] is not None:
                try:
                    value = float(data[field])
                    if value < 0:
                        errors.append(f"{field} cannot be negative: {value}")
                    
                    # Specific range checks
                    if field in ['scrap_content_percentage', 'recycling_rate_percentage'] and value > 100:
                        errors.append(f"{field} cannot exceed 100%: {value}")
                        
                except (ValueError, TypeError):
                    errors.append(f"{field} must be numeric: {data[field]}")
        
        return errors
    
    @classmethod
    def process_excel_file(cls, file_path: str, project_id: str, user_id: int) -> Tuple[List[Dict], List[str], List[str]]:
        """
        Process Excel/CSV file and prepare batches for database
        
        Returns:
            Tuple of (batches, warnings, errors)
        """
        batches = []
        warnings = []
        errors = []
        
        try:
            # Read file
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:  # Excel
                df = pd.read_excel(file_path)
            
            # Clean column names
            df.columns = [str(col).strip() for col in df.columns]
            
            # Map columns
            column_mapping = cls.map_columns(df.columns.tolist())
            
            if not column_mapping:
                errors.append("Could not map any columns to database fields. Please check file format.")
                return batches, warnings, errors
            
            # Warn about unmapped columns
            unmapped = [col for col in df.columns if col not in column_mapping]
            if unmapped:
                warnings.append(f"Unmapped columns (ignored): {', '.join(unmapped)}")
            
            # Process each row
            for idx, row in df.iterrows():
                try:
                    # Build batch data from mapped columns
                    batch_data = {
                        "project_id": project_id
                    }
                    
                    for file_col, db_field in column_mapping.items():
                        if file_col in row:
                            value = row[file_col]
                            # Handle NaN values
                            if pd.isna(value):
                                batch_data[db_field] = None
                            else:
                                batch_data[db_field] = value
                    
                    # Ensure batch_id exists
                    if 'batch_id' not in batch_data or not batch_data['batch_id']:
                        batch_data['batch_id'] = f"import_{project_id}_{datetime.now().strftime('%Y%m%d')}_{idx+1}"
                    
                    # Set default values for missing optional fields
                    optional_fields = {
                        'raw_material_type': 'Unknown',
                        'material_type': 'Steel',
                        'energy_source_type': 'Electricity',
                        'recycled_material_quantity': 0,
                        'water_consumption_liters': 0,
                        'ore_grade': 0.5,
                        'waste_slag_quantity': 0,
                        'scrap_content_percentage': 0,
                        'recycling_rate_percentage': 0
                    }
                    
                    for field, default_value in optional_fields.items():
                        if field not in batch_data or batch_data[field] is None:
                            batch_data[field] = default_value
                    
                    # Validate
                    validation_errors = cls.validate_batch_data(batch_data)
                    
                    if validation_errors:
                        errors.append(f"Row {idx+1} ({batch_data.get('batch_id', 'No ID')}): {', '.join(validation_errors)}")
                        continue
                    
                    # FIXED: Get full prediction with uncertainty
                    prediction_payload = {
                        "raw_material_type": batch_data["raw_material_type"],
                        "material_type": batch_data["material_type"],
                        "energy_source_type": batch_data["energy_source_type"],
                        "raw_material_quantity": float(batch_data["raw_material_quantity"]),
                        "recycled_material_quantity": float(batch_data.get("recycled_material_quantity", 0)),
                        "energy_consumption_kwh": float(batch_data["energy_consumption_kwh"]),
                        "water_consumption_liters": float(batch_data.get("water_consumption_liters", 0)),
                        "production_volume": float(batch_data["production_volume"]),
                        "ore_grade": float(batch_data.get("ore_grade", 0.5)),
                        "waste_slag_quantity": float(batch_data.get("waste_slag_quantity", 0)),
                        "scrap_content_percentage": float(batch_data.get("scrap_content_percentage", 0)),
                        "recycling_rate_percentage": float(batch_data.get("recycling_rate_percentage", 0))
                    }
                    
                    prediction_result = predict_emissions_with_uncertainty(
                        prediction_payload, 
                        include_explanations=False
                    )
                    
                    co2 = prediction_result["predictions"]["co2_emissions_kg"]
                    sox = prediction_result["predictions"]["sox_emissions_kg"]
                    nox = prediction_result["predictions"]["nox_emissions_kg"]
                    uncertainty = prediction_result["uncertainty"]
                    model_metadata = prediction_result["model_metadata"]
                    
                    # Calculate derived metrics
                    production_volume = float(batch_data["production_volume"])
                    if production_volume > 0:
                        carbon_intensity = co2 / production_volume
                        energy_intensity = float(batch_data["energy_consumption_kwh"]) / production_volume
                        water_intensity = float(batch_data.get("water_consumption_liters", 0)) / production_volume
                        recycling_efficiency = (
                            float(batch_data.get("recycled_material_quantity", 0)) / 
                            float(batch_data["raw_material_quantity"]) * 100
                            if float(batch_data["raw_material_quantity"]) > 0 else 0
                        )
                    else:
                        carbon_intensity = 0
                        energy_intensity = 0
                        water_intensity = 0
                        recycling_efficiency = 0
                    
                    # FIXED: Add ML outputs and derived metrics with real uncertainty data
                    batch_data.update({
                        "co2_emissions_kg": float(co2),
                        "sox_emissions_kg": float(sox),
                        "nox_emissions_kg": float(nox),
                        "carbon_intensity_kg_per_ton": float(carbon_intensity),
                        "energy_intensity_kwh_per_ton": float(energy_intensity),
                        "water_intensity_l_per_ton": float(water_intensity),
                        "recycling_efficiency_score": float(recycling_efficiency),
                        "co2_lower_bound": float(uncertainty["co2"]["lower_bound"]),
                        "co2_upper_bound": float(uncertainty["co2"]["upper_bound"]),
                        "sox_lower_bound": float(uncertainty["sox"]["lower_bound"]),
                        "sox_upper_bound": float(uncertainty["sox"]["upper_bound"]),
                        "nox_lower_bound": float(uncertainty["nox"]["lower_bound"]),
                        "nox_upper_bound": float(uncertainty["nox"]["upper_bound"]),
                        "prediction_confidence": float(model_metadata["prediction_confidence"]),
                        "uncertainty_data": uncertainty,
                        "model_metadata": model_metadata
                    })
                    
                    batches.append(batch_data)
                    
                except Exception as e:
                    errors.append(f"Row {idx+1} processing error: {str(e)}")
            
            # Summary
            if batches:
                warnings.append(f"Successfully processed {len(batches)} rows")
            
        except Exception as e:
            errors.append(f"File processing error: {str(e)}")
        
        return batches, warnings, errors
    
    @classmethod
    def generate_template_excel(cls) -> bytes:
        """Generate Excel template for data upload"""
        import io
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Process Data Template"
        
        # Headers with descriptions
        headers = [
            ("Batch ID", "Unique identifier for the batch"),
            ("Raw Material Type", "e.g., Iron Ore, Bauxite, Copper Ore"),
            ("Material Type", "e.g., Steel, Aluminium, Copper"),
            ("Energy Source Type", "e.g., Electricity, Coal, Natural Gas"),
            ("Raw Material Quantity", "kg - Virgin material input"),
            ("Recycled Material Quantity", "kg - Recycled/scrap material input"),
            ("Energy Consumption", "kWh - Total energy used"),
            ("Water Consumption", "Liters - Total water used"),
            ("Production Volume", "kg - Total output produced"),
            ("Ore Grade", "0-1 - Quality of raw material (e.g., 0.65 for 65%)"),
            ("Waste Slag Quantity", "kg - Waste generated"),
            ("Scrap Content %", "0-100 - Percentage of scrap in input"),
            ("Recycling Rate %", "0-100 - Percentage of material recycled")
        ]
        
        # Write headers
        for col, (header, description) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)
            
            # Add description as comment
            from openpyxl.comments import Comment
            cell.comment = Comment(description, "EcoMetal LCA")
        
        # Add example data
        example_row = [
            "BATCH_001",
            "Iron Ore",
            "Steel",
            "Electricity",
            1500,
            300,
            4500,
            2500,
            1200,
            0.65,
            150,
            20,
            25
        ]
        
        for col, value in enumerate(example_row, start=1):
            ws.cell(row=2, column=col, value=value)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()